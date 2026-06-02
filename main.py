import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
from copy import copy


def resource_path(name: str) -> str:
    """Works both from source and inside a PyInstaller --onefile bundle."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False

from openpyxl import load_workbook, Workbook

BG = "#f0f2f5"
ACCENT = "#2563eb"
TEXT = "#1e293b"
MUTED = "#64748b"
DROP_BG = "#e8f0fe"
DROP_BORDER = "#93c5fd"
CARD_BG = "#ffffff"
BTN_SECONDARY = "#475569"


# ── xlsx formatting helpers ───────────────────────────────────────────────────

def _copy_cell(src, dest) -> None:
    dest.value = src.value
    if src.has_style:
        dest.font = copy(src.font)
        dest.fill = copy(src.fill)
        dest.border = copy(src.border)
        dest.alignment = copy(src.alignment)
        dest.number_format = src.number_format
        dest.protection = copy(src.protection)


def _copy_col_dims(src_ws, dest_ws) -> None:
    for col, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col].width = dim.width
        dest_ws.column_dimensions[col].hidden = dim.hidden


def _copy_row_height(src_ws, src_idx: int, dest_ws, dest_idx: int) -> None:
    dim = src_ws.row_dimensions.get(src_idx)
    if dim:
        dest_ws.row_dimensions[dest_idx].height = dim.height
        dest_ws.row_dimensions[dest_idx].hidden = dim.hidden


# ── chunk logic ───────────────────────────────────────────────────────────────

def _make_chunks(data: list, mode: str, value: int) -> list[list]:
    """Split data rows without dropping any row.

    mode='parts': exactly `value` parts, distributed as evenly as possible.
    mode='rows':  each part has at most `value` rows.
    """
    if not data:
        return []

    if mode == "rows":
        return [data[i: i + value] for i in range(0, len(data), value)]

    # mode == "parts"
    n = min(value, len(data))
    base, extra = divmod(len(data), n)
    chunks, start = [], 0
    for i in range(n):
        size = base + (1 if i < extra else 0)
        chunks.append(data[start: start + size])
        start += size
    return chunks


# ── app ───────────────────────────────────────────────────────────────────────

_Base = TkinterDnD.Tk if HAS_DND else tk.Tk


class App(_Base):
    def __init__(self):
        super().__init__()
        self.title("Табличник")
        self.geometry("640x540")
        self.minsize(480, 420)
        self.configure(bg=BG)
        self.files: list[str] = []
        ico = resource_path("icon.ico")
        if os.path.exists(ico):
            self.iconbitmap(ico)
        self._build_ui()
        if HAS_DND:
            self._setup_dnd()

    def _build_ui(self):
        header = tk.Frame(self, bg=BG, pady=14)
        header.pack(fill=tk.X, padx=24)
        tk.Label(header, text="Табличник", font=("Segoe UI", 20, "bold"),
                 bg=BG, fg=TEXT).pack(side=tk.LEFT)
        tk.Label(header, text="  работа с .xlsx файлами",
                 font=("Segoe UI", 12), bg=BG, fg=MUTED).pack(side=tk.LEFT, pady=4)

        self.drop_frame = tk.Frame(
            self, bg=DROP_BG, bd=0,
            highlightthickness=2,
            highlightbackground=DROP_BORDER,
            highlightcolor=ACCENT)
        self.drop_frame.pack(fill=tk.X, padx=24, pady=(0, 12))

        self.drop_label = tk.Label(
            self.drop_frame,
            text="⬇   Перетащите .xlsx файлы сюда",
            font=("Segoe UI", 13), bg=DROP_BG, fg=MUTED,
            pady=26, padx=20)
        self.drop_label.pack(fill=tk.X)

        tk.Button(
            self.drop_frame, text="Выбрать файлы вручную",
            font=("Segoe UI", 10), bg=DROP_BG, fg=ACCENT,
            relief=tk.FLAT, cursor="hand2",
            command=self._pick_files
        ).pack(pady=(0, 10))

        list_outer = tk.Frame(self, bg=BG)
        list_outer.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 12))

        tk.Label(list_outer, text="Файлы:", font=("Segoe UI", 10, "bold"),
                 bg=BG, fg=TEXT).pack(anchor=tk.W, pady=(0, 4))

        card = tk.Frame(list_outer, bg=CARD_BG, bd=0,
                        highlightthickness=1, highlightbackground="#e2e8f0")
        card.pack(fill=tk.BOTH, expand=True)

        self.listbox = tk.Listbox(
            card, font=("Segoe UI", 10), bg=CARD_BG, fg=TEXT,
            selectbackground=ACCENT, selectforeground="white",
            relief=tk.FLAT, bd=0, activestyle="none",
            selectmode=tk.EXTENDED)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=6)

        sb = ttk.Scrollbar(card, command=self.listbox.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=sb.set)

        tk.Button(
            list_outer, text="Удалить выбранные",
            font=("Segoe UI", 9), bg=BG, fg=MUTED,
            relief=tk.FLAT, cursor="hand2",
            command=self._remove_selected
        ).pack(anchor=tk.E, pady=(4, 0))

        btn_row = tk.Frame(self, bg=BG)
        btn_row.pack(fill=tk.X, padx=24, pady=(0, 18))

        self.btn_merge = tk.Button(
            btn_row, text="Объединить  ▶",
            font=("Segoe UI", 11, "bold"), bg=ACCENT, fg="white",
            relief=tk.FLAT, padx=20, pady=8, cursor="hand2",
            command=self._merge, state=tk.DISABLED)
        self.btn_merge.pack(side=tk.LEFT, padx=(0, 8))

        self.btn_split = tk.Button(
            btn_row, text="Разделить  ▶",
            font=("Segoe UI", 11, "bold"), bg=BTN_SECONDARY, fg="white",
            relief=tk.FLAT, padx=20, pady=8, cursor="hand2",
            command=self._split, state=tk.DISABLED)
        self.btn_split.pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(
            btn_row, text="Очистить",
            font=("Segoe UI", 10), bg=BG, fg=MUTED,
            relief=tk.FLAT, padx=12, pady=8, cursor="hand2",
            command=self._clear
        ).pack(side=tk.RIGHT)

        self.status_var = tk.StringVar(value="Готов")
        tk.Label(self, textvariable=self.status_var,
                 font=("Segoe UI", 9), bg="#e2e8f0", fg=MUTED,
                 anchor=tk.W, padx=12, pady=4).pack(fill=tk.X, side=tk.BOTTOM)

    def _setup_dnd(self):
        for widget in (self.drop_frame, self.drop_label):
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", self._on_drop)
            widget.dnd_bind("<<DragEnter>>", self._on_drag_enter)
            widget.dnd_bind("<<DragLeave>>", self._on_drag_leave)

    def _on_drag_enter(self, _event):
        self.drop_frame.config(highlightbackground=ACCENT)
        self.drop_label.config(fg=ACCENT)

    def _on_drag_leave(self, _event):
        self.drop_frame.config(highlightbackground=DROP_BORDER)
        self.drop_label.config(fg=MUTED)

    def _on_drop(self, event):
        self._on_drag_leave(None)
        paths = self.tk.splitlist(event.data)
        added = [p for p in paths if p.lower().endswith(".xlsx") and p not in self.files]
        for p in added:
            self.files.append(p)
            self.listbox.insert(tk.END, os.path.basename(p))
        self._update_buttons()
        if not added:
            return
        self.status_var.set(f"Добавлено: {len(added)}. Всего файлов: {len(self.files)}")
        if len(self.files) >= 2:
            if messagebox.askyesno(
                    "Объединить?",
                    f"Добавлено {len(self.files)} файла(ов).\nОбъединить их в один?"):
                self._merge()
        elif len(self.files) == 1:
            if messagebox.askyesno(
                    "Разделить?",
                    "Добавлен 1 файл.\nРазделить его на несколько частей?"):
                self._split()

    def _pick_files(self):
        paths = filedialog.askopenfilenames(
            title="Выберите .xlsx файлы",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        added = 0
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.listbox.insert(tk.END, os.path.basename(p))
                added += 1
        self._update_buttons()
        if added:
            self.status_var.set(f"Добавлено: {added}. Всего файлов: {len(self.files)}")

    def _remove_selected(self):
        for i in reversed(self.listbox.curselection()):
            self.listbox.delete(i)
            self.files.pop(i)
        self._update_buttons()
        self.status_var.set(f"Файлов: {len(self.files)}")

    def _clear(self):
        self.files.clear()
        self.listbox.delete(0, tk.END)
        self._update_buttons()
        self.status_var.set("Очищено")

    def _update_buttons(self):
        n = len(self.files)
        self.btn_merge.config(state=tk.NORMAL if n >= 2 else tk.DISABLED)
        self.btn_split.config(state=tk.NORMAL if n >= 1 else tk.DISABLED)

    # ── Merge ────────────────────────────────────────────────────────────────

    def _merge(self):
        if len(self.files) < 2:
            messagebox.showwarning("Нужно минимум 2 файла", "Добавьте хотя бы 2 .xlsx файла.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Сохранить объединённый файл",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="merged.xlsx")
        if not out_path:
            return

        try:
            self.status_var.set("Объединяю…")
            self.update_idletasks()

            out_wb = Workbook()
            out_ws = out_wb.active
            dest_row = 1
            first = True

            for fpath in self.files:
                wb = load_workbook(fpath)
                ws = wb.active
                src_rows = list(ws.iter_rows())
                if not src_rows:
                    continue

                if first:
                    _copy_col_dims(ws, out_ws)
                    first = False
                    start = 0          # include header
                else:
                    start = 1          # skip header of subsequent files

                for src_idx, row in enumerate(src_rows[start:], start + 1):
                    for col_idx, src_cell in enumerate(row, 1):
                        _copy_cell(src_cell, out_ws.cell(row=dest_row, column=col_idx))
                    _copy_row_height(ws, src_idx, out_ws, dest_row)
                    dest_row += 1

            out_wb.save(out_path)
            self.status_var.set(
                f"Объединено {len(self.files)} файлов → {os.path.basename(out_path)}")
            messagebox.showinfo("Готово!", f"Файлы объединены:\n{out_path}")
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            self.status_var.set("Ошибка при объединении")

    # ── Split ────────────────────────────────────────────────────────────────

    def _split(self):
        if len(self.files) != 1:
            messagebox.showwarning(
                "Нужен 1 файл",
                "Для разделения выберите ровно один файл.\n"
                "Для нескольких — сначала удалите лишние.")
            return

        dlg = SplitDialog(self)
        self.wait_window(dlg)
        if not dlg.result:
            return

        mode, value = dlg.result

        out_dir = filedialog.askdirectory(title="Папка для сохранения частей")
        if not out_dir:
            return

        fpath = self.files[0]
        try:
            self.status_var.set("Разделяю…")
            self.update_idletasks()

            wb = load_workbook(fpath)
            ws = wb.active
            src_rows = list(ws.iter_rows())
            if not src_rows:
                messagebox.showwarning("Файл пуст", "Файл не содержит данных.")
                return

            header_row = src_rows[0]
            data_rows = src_rows[1:]
            chunks = _make_chunks(data_rows, mode, value)

            base = os.path.splitext(os.path.basename(fpath))[0]
            saved = 0
            for i, chunk in enumerate(chunks):
                new_wb = Workbook()
                new_ws = new_wb.active

                _copy_col_dims(ws, new_ws)

                # header → row 1
                for col_idx, src_cell in enumerate(header_row, 1):
                    _copy_cell(src_cell, new_ws.cell(row=1, column=col_idx))
                _copy_row_height(ws, 1, new_ws, 1)

                # data → rows 2+
                for dest_idx, src_row in enumerate(chunk, 2):
                    src_row_idx = src_row[0].row
                    for col_idx, src_cell in enumerate(src_row, 1):
                        _copy_cell(src_cell, new_ws.cell(row=dest_idx, column=col_idx))
                    _copy_row_height(ws, src_row_idx, new_ws, dest_idx)

                out_path = os.path.join(out_dir, f"{base}_часть{i + 1}.xlsx")
                new_wb.save(out_path)
                saved += 1

            self.status_var.set(f"Разделено на {saved} частей → {out_dir}")
            messagebox.showinfo("Готово!", f"Файл разделён на {saved} частей.\nПапка: {out_dir}")
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            self.status_var.set("Ошибка при разделении")


class SplitDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Параметры разделения")
        self.geometry("330x210")
        self.resizable(False, False)
        self.configure(bg=BG)
        self.result = None
        self.grab_set()
        self.transient(parent)
        self._build()

    def _build(self):
        tk.Label(self, text="Как разделить файл?",
                 font=("Segoe UI", 12, "bold"), bg=BG, fg=TEXT,
                 pady=16).pack()

        self.mode = tk.StringVar(value="parts")

        row1 = tk.Frame(self, bg=BG)
        row1.pack(fill=tk.X, padx=28, pady=4)
        tk.Radiobutton(row1, text="На", variable=self.mode, value="parts",
                       bg=BG, font=("Segoe UI", 10),
                       command=self._toggle).pack(side=tk.LEFT)
        self.parts_spin = tk.Spinbox(row1, from_=2, to=10000, width=6,
                                     font=("Segoe UI", 10))
        self.parts_spin.pack(side=tk.LEFT, padx=6)
        tk.Label(row1, text="равных частей", bg=BG,
                 font=("Segoe UI", 10), fg=TEXT).pack(side=tk.LEFT)

        row2 = tk.Frame(self, bg=BG)
        row2.pack(fill=tk.X, padx=28, pady=4)
        tk.Radiobutton(row2, text="По", variable=self.mode, value="rows",
                       bg=BG, font=("Segoe UI", 10),
                       command=self._toggle).pack(side=tk.LEFT)
        self.rows_spin = tk.Spinbox(row2, from_=1, to=1_000_000, width=8,
                                    font=("Segoe UI", 10), state=tk.DISABLED)
        self.rows_spin.pack(side=tk.LEFT, padx=6)
        tk.Label(row2, text="строк на файл", bg=BG,
                 font=("Segoe UI", 10), fg=TEXT).pack(side=tk.LEFT)

        btns = tk.Frame(self, bg=BG)
        btns.pack(pady=18)
        tk.Button(btns, text="Разделить",
                  font=("Segoe UI", 10, "bold"), bg=ACCENT, fg="white",
                  relief=tk.FLAT, padx=16, pady=6, cursor="hand2",
                  command=self._ok).pack(side=tk.LEFT, padx=4)
        tk.Button(btns, text="Отмена",
                  font=("Segoe UI", 10), bg="#e2e8f0", fg=TEXT,
                  relief=tk.FLAT, padx=16, pady=6, cursor="hand2",
                  command=self.destroy).pack(side=tk.LEFT, padx=4)

    def _toggle(self):
        if self.mode.get() == "parts":
            self.parts_spin.config(state=tk.NORMAL)
            self.rows_spin.config(state=tk.DISABLED)
        else:
            self.parts_spin.config(state=tk.DISABLED)
            self.rows_spin.config(state=tk.NORMAL)

    def _ok(self):
        try:
            if self.mode.get() == "parts":
                v = int(self.parts_spin.get())
                if v < 2:
                    messagebox.showwarning("Ошибка", "Минимум 2 части.", parent=self)
                    return
                self.result = ("parts", v)
            else:
                v = int(self.rows_spin.get())
                if v < 1:
                    messagebox.showwarning("Ошибка", "Минимум 1 строка.", parent=self)
                    return
                self.result = ("rows", v)
            self.destroy()
        except ValueError:
            messagebox.showwarning("Ошибка", "Введите корректное число.", parent=self)


if __name__ == "__main__":
    app = App()
    app.mainloop()
